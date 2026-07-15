from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
from brand import *
import sample_data as S

wb=Workbook()
def D(s): return datetime.strptime(s,"%Y-%m-%d")
CLI={c[0]:c[1] for c in S.CLIENTES}
PACN={p[0]:p[1] for p in S.PACOTES}; PACT={p[0]:p[2] for p in S.PACOTES}; PACD={p[0]:p[3] for p in S.PACOTES}
FORN={x[0]:x[1] for x in S.FORNECEDORES}

# ===== INÍCIO =====
ws=wb.active; ws.title="Início"; nogrid(ws); tab(ws,BLUE_DK)
ws["B2"]="Let's Go ✈ — Financeiro"; ws["B2"].font=Font(name=FONT,bold=True,size=22,color=BLUE)
ws["B3"]="Confidencial · só a gestão 🔒"; ws["B3"].font=f(11,color=PINK)
def sec(row,txt,color):
    c=ws.cell(row=row,column=2,value=txt); c.font=Font(name=FONT,bold=True,size=11,color="FFFFFF")
    c.fill=fill(color); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=8); ws.row_dimensions[row].height=22
def tx(row,t): ws.cell(row=row,column=2,value=t).font=f(10,color=TEXT)
r=5
sec(r,"  Ligação ao operacional",BLUE); r+=1
for t in ["A aba «Importar Reservas» recebe o que a equipa insere no ficheiro operacional.",
          "No Google Sheets, cola na célula A3 dessa aba:",
          '     =IMPORTRANGE("URL_DO_FICHEIRO_OPERACIONAL"; "Reservas!A2:AK")',
          "Na 1.ª vez o Sheets pede «Permitir acesso». É uma ligação de sentido único (só lê)."]:
    tx(r,"•  "+t); r+=1
r+=1
sec(r,"  O que este ficheiro faz 💰",ORANGE); r+=1
for t in ["Comissões por funcionário (confidencial) · Custos fixos · Fluxo de caixa.",
          "Contas a receber · Lucros (P&L) · Dashboard. O custo usado é o custo TOTAL de fornecedores da reserva."]:
    tx(r,"•  "+t); r+=1
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=120

# ===== CONFIGURAÇÕES (financeiro) =====
ws=wb.create_sheet("Configurações"); nogrid(ws); tab(ws,MUTE)
ws["A1"]="⚙️ Configurações (confidencial)"; ws["A1"].font=Font(name=FONT,bold=True,size=15,color=BLUE)
ws["A2"]="Regras de comissão, câmbios e definições. Mudar aqui altera os cálculos sem mexer em fórmulas."; ws["A2"].font=f(9,it=True,color=MUTE)
ws["A3"]="Comissões por funcionário"; ws["A3"].font=f(11,b=True,color=PINK)
paint_header(ws,4,["Funcionário","Taxa (%)","Base de cálculo"],[16,12,18],fillhex=ORANGE)
regras=[["Sofia",0.05,"Margem"],["Miguel",0.04,"Margem"]]
for i,rw in enumerate(regras,5):
    ws.cell(row=i,column=1,value=rw[0]).font=f()
    c=ws.cell(row=i,column=2,value=rw[1]); c.number_format=PCT; c.fill=fill(ORANGE_LT); c.font=f(b=True,color=BLUE_DK); c.border=BOT
    c=ws.cell(row=i,column=3,value=rw[2]); c.fill=fill(BLUE_LT); c.font=f(color=BLUE_DK); c.border=BOT
    ws.cell(row=i,column=1).border=BOT
ws.cell(row=5+len(regras)+1,column=1,value="Base «Margem» = venda − custo total dos fornecedores.").font=f(8,it=True,color=MUTE)
# taxas / moedas
ws["E3"]="Câmbios (→ EUR)"; ws["E3"].font=f(11,b=True,color=PINK)
paint_header(ws,4,["Moeda","Câmbio p/ EUR"],[12,16],fillhex=BLUE_MD)  # note: paint at cols A? -> override below
# escrever manualmente para não colidir com cols A-C
ws.cell(row=4,column=5,value="Moeda").font=Font(name=FONT,bold=True,size=10,color="FFFFFF"); ws.cell(row=4,column=5).fill=fill(BLUE_MD); ws.cell(row=4,column=5).alignment=Alignment(horizontal="center")
ws.cell(row=4,column=6,value="Câmbio p/ EUR").font=Font(name=FONT,bold=True,size=10,color="FFFFFF"); ws.cell(row=4,column=6).fill=fill(BLUE_MD); ws.cell(row=4,column=6).alignment=Alignment(horizontal="center")
moedas=[["EUR",1.00],["USD",0.92],["GBP",1.17],["BRL",0.18]]
for i,rw in enumerate(moedas,5):
    ws.cell(row=i,column=5,value=rw[0]).font=f(); c=ws.cell(row=i,column=6,value=rw[1]); c.number_format='0.0000'; c.fill=fill(BLUE_LT); c.font=f(color=BLUE_DK)
    for col in (5,6): ws.cell(row=i,column=col).border=BOT
ws.cell(row=5+len(moedas)+1,column=5,value="Atualizar à mão quando necessário.").font=f(8,it=True,color=MUTE)
# definições
ws["A11"]="Definições / notas"; ws["A11"].font=f(11,b=True,color=PINK)
notas=["Regime de IVA: confirmar com a contabilista (agências costumam usar o regime de margem).",
       "Custos fixos: ver aba «Custos Fixos».  Gastos avulsos: ver aba «Gastos Adicionais»."]
for k,t in enumerate(notas,12): ws.cell(row=k,column=1,value="•  "+t).font=f(9,color=TEXT)
ws.column_dimensions["A"].width=18; ws.column_dimensions["B"].width=12; ws.column_dimensions["C"].width=18
ws.column_dimensions["D"].width=3; ws.column_dimensions["E"].width=12; ws.column_dimensions["F"].width=16
RGF=5; RGL=4+len(regras)
for nm,ref in [("rngComFunc",f"'Configurações'!$A${RGF}:$A${RGF+9}"),
               ("rngComTaxa",f"'Configurações'!$B${RGF}:$B${RGF+9}"),
               ("rngComBase",f"'Configurações'!$C${RGF}:$C${RGF+9}")]:
    wb.defined_names.add(DefinedName(nm,attr_text=ref))

# ===== IMPORTAR RESERVAS =====
ws=wb.create_sheet("Importar Reservas"); nogrid(ws); tab(ws,BLUE_MD)
note=ws.cell(row=1,column=1,value='Dados de exemplo — substituir por  =IMPORTRANGE("URL_operacional";"Reservas!A2:AF")  no Google Sheets')
note.font=f(9,b=True,color="9C6500"); note.fill=fill(ORANGE_LT); ws.merge_cells("A1:AF1")
H=["ID Reserva","Data registo","ID Cliente","Cliente","ID Pacote","Pacote","Tipo","Destino","Data início","Data fim","Nº pax","Estado",
   "Estado aprovação","Custo total (€)","Valor venda (€)","Margem (€)","Aprovado por",
   "Tipo pagamento","Nº prestações","Prestações pagas","Valor/prestação (€)","Valor pago (€)","Valor pendente (€)","Prestações em falta","% pago",
   "Vendedor","Criado por","Motivo","Contactado?","Próxima ação","Data follow-up","Observações"]
paint_header(ws,2,H,[13,11,10,15,10,18,10,14,11,11,7,15,16,14,14,13,13,13,12,13,15,14,15,13,8,14,14,22,11,24,13,22],fillhex=BLUE_MD)
row=3
for r in S.RESERVAS:
    (idr,dreg,idc,idp,dini,dfim,pax,estado,aprov,venda,tpag,nprest,pagas,func,motivo,contactado,prox,dfu,obs)=r
    dv=S.derived(r); vv=venda if isinstance(venda,(int,float)) else ""
    marg=dv['margem'] if dv['margem']!="" else ""
    up=lambda x: x if vv!="" else ""
    vals=[idr,D(dreg),idc,CLI.get(idc,""),idp,PACN.get(idp,""),PACT.get(idp,""),PACD.get(idp,""),D(dini),D(dfim),pax,estado,
          aprov,dv['total'],vv,marg,("Gestora" if aprov=="Aprovado" else ""),
          tpag,nprest,pagas,up(dv['vprest']),up(dv['pago']),up(dv['pend']),dv['falta'],up(dv['pct']),
          func,func,motivo,contactado,prox,(D(dfu) if dfu else ""),obs]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=row,column=j,value=v)
        if j in (2,9,10,31): c.number_format=DATE
        if j in (14,15,16,21,22,23): c.number_format=EUR
        if j==25: c.number_format=PCT
    row+=1
zebra(ws,3,row-1,len(H)); ws.freeze_panes="A3"

# ===== COMISSÕES =====
ws=wb.create_sheet("Comissões"); nogrid(ws); tab(ws,ORANGE)
ws["A1"]="💼 Comissões (confidencial)"; ws["A1"].font=Font(name=FONT,bold=True,size=15,color=BLUE)
ws["A3"]="Regras (taxas) definidas na aba «Configurações»."; ws["A3"].font=f(10,it=True,color=MUTE)
regras=[["Sofia",0.05,"Margem"],["Miguel",0.04,"Margem"]]
ws["A9"]="Cálculo por reserva (puxa de «Importar Reservas»)"; ws["A9"].font=f(11,b=True,color=PINK)
H=["ID Reserva","Data","Funcionário","Valor venda (€)","Custo total (€)","Base","Taxa","Comissão (€)"]; W=[13,12,14,15,15,10,8,14]
paint_header(ws,10,H,W)
r0=11; LAST=42; IR="'Importar Reservas'"
for k in range(LAST-2):
    rr=r0+k; src=3+k; g=f"{IR}!$A{src}=\"\""
    ws.cell(row=rr,column=1,value=f'=IF({g},"",{IR}!$A{src})')
    c=ws.cell(row=rr,column=2,value=f'=IF({g},"",{IR}!$B{src})'); c.number_format=DATE
    ws.cell(row=rr,column=3,value=f'=IF({g},"",{IR}!$Z{src})')
    c=ws.cell(row=rr,column=4,value=f'=IF({g},"",{IR}!$O{src})'); c.number_format=EUR
    c=ws.cell(row=rr,column=5,value=f'=IF({g},"",{IR}!$N{src})'); c.number_format=EUR
    ws.cell(row=rr,column=6,value=f'=IFERROR(IF($C{rr}="","",INDEX(rngComBase,MATCH($C{rr},rngComFunc,0))),"")')
    c=ws.cell(row=rr,column=7,value=f'=IFERROR(IF($C{rr}="","",INDEX(rngComTaxa,MATCH($C{rr},rngComFunc,0))),"")'); c.number_format=PCT
    c=ws.cell(row=rr,column=8,value=f'=IF(OR($C{rr}="",$D{rr}=""),"",IF($F{rr}="Margem",($D{rr}-$E{rr}),$D{rr})*$G{rr})'); c.number_format=EUR
    for col in range(1,9): ws.cell(row=rr,column=col).border=BOT; ws.cell(row=rr,column=col).font=f()
    ws.cell(row=rr,column=2).number_format=DATE; ws.cell(row=rr,column=4).number_format=EUR
    ws.cell(row=rr,column=5).number_format=EUR; ws.cell(row=rr,column=7).number_format=PCT; ws.cell(row=rr,column=8).number_format=EUR
rN=r0+(LAST-2)-1
ws["J10"]="Total a pagar"; ws["J10"].font=f(11,b=True,color=PINK)
for cc,t in [(10,"Funcionário"),(11,"Comissão (€)")]:
    c=ws.cell(row=11,column=cc,value=t); c.font=Font(name=FONT,bold=True,color="FFFFFF",size=10); c.fill=fill(ORANGE); c.alignment=Alignment(horizontal="center")
ws.column_dimensions["J"].width=16; ws.column_dimensions["K"].width=14
for i,rw in enumerate(regras,12):
    ws.cell(row=i,column=10,value=rw[0]).font=f(); ws.cell(row=i,column=10).border=BOT
    c=ws.cell(row=i,column=11,value=f'=SUMIF($C${r0}:$C${rN},$J{i},$H${r0}:$H${rN})'); c.number_format=EUR; c.border=BOT
TOT0,TOT1=12,11+len(regras)
ws.cell(row=TOT1+1,column=10,value="TOTAL").font=f(b=True)
c=ws.cell(row=TOT1+1,column=11,value=f'=SUM($K${TOT0}:$K${TOT1})'); c.number_format=EUR; c.font=f(b=True,color=ORANGE)
COMM_TOTAL=f"Comissões!$K${TOT1+1}"

# ===== CUSTOS FIXOS =====
ws=wb.create_sheet("Custos Fixos"); nogrid(ws); tab(ws,PINK)
ws["A1"]="🏢 Custos fixos"; ws["A1"].font=Font(name=FONT,bold=True,size=15,color=BLUE)
H=["Descrição","Categoria","Valor (€)","Periodicidade","Início","Fim","Valor mensal equiv. (€)"]; W=[24,16,14,15,12,12,20]
paint_header(ws,2,H,W,fillhex=PINK)
custos=[["Renda do escritório","Instalações",650,"Mensal","2026-01-01",""],["Salário — Sofia","Salários",900,"Mensal","2026-01-01",""],
 ["Salário — Miguel","Salários",850,"Mensal","2026-01-01",""],["Software de gestão","Software",30,"Mensal","2026-01-01",""],
 ["Google Workspace","Software",12,"Mensal","2026-01-01",""],["Contabilidade","Serviços",120,"Mensal","2026-01-01",""],
 ["Seguro anual","Serviços",360,"Anual","2026-01-01",""]]
r0=3
for i,rw in enumerate(custos,r0):
    ws.cell(row=i,column=1,value=rw[0]).font=f(); ws.cell(row=i,column=2,value=rw[1]).font=f()
    c=ws.cell(row=i,column=3,value=rw[2]); c.number_format=EUR; c.fill=fill(BLUE_LT); c.font=f(color=BLUE_DK)
    ws.cell(row=i,column=4,value=rw[3]).font=f()
    c=ws.cell(row=i,column=5)
    if rw[4]: c.value=D(rw[4]); c.number_format=DATE
    ws.cell(row=i,column=6,value=rw[5])
    c=ws.cell(row=i,column=7,value=f'=IF($D{i}="Mensal",$C{i},IF($D{i}="Trimestral",$C{i}/3,IF($D{i}="Anual",$C{i}/12,$C{i})))'); c.number_format=EUR
    for col in range(1,8): ws.cell(row=i,column=col).border=BOT
rN=r0+len(custos)-1
ws.cell(row=rN+1,column=6,value="TOTAL MENSAL").font=f(b=True); ws.cell(row=rN+1,column=6).alignment=Alignment(horizontal="right")
c=ws.cell(row=rN+1,column=7,value=f'=SUM($G${r0}:$G${rN})'); c.number_format=EUR; c.font=f(b=True,color=BLUE)
CF=f"'Custos Fixos'!$G${rN+1}"

# ===== GASTOS ADICIONAIS (input da gestora) =====
ws=wb.create_sheet("Gastos Adicionais"); nogrid(ws); tab(ws,PINK)
ws["A1"]="🧾 Gastos adicionais"; ws["A1"].font=Font(name=FONT,bold=True,size=15,color=BLUE)
ws["A2"]="Despesas fora dos custos fixos e dos fornecedores. Escreve aqui — vão sozinhas para o Fluxo de Caixa e para o P&L (Lucro)."
ws["A2"].font=f(9,it=True,color=MUTE)
GH=["Data","Categoria","Descrição","Valor (€)","Método","Observações"]; GW=[13,18,30,14,16,26]
paint_header(ws,3,GH,GW,fillhex=PINK)
gastos=[["2026-02-10","Marketing","Anúncios Facebook/Instagram",150,"MB Way",""],
 ["2026-02-25","Material escritório","Toner + papel",60,"Cartão",""],
 ["2026-03-08","Deslocações","Feira de turismo (combustível/portagens)",85,"Numerário",""],
 ["2026-03-20","Taxas/bancos","Comissões bancárias",25,"Débito direto",""],
 ["2026-04-05","Equipamento","Impressora nova",180,"Transferência",""]]
GR0=4
for i,rw in enumerate(gastos,GR0):
    c=ws.cell(row=i,column=1,value=D(rw[0])); c.number_format=DATE; c.font=f()
    ws.cell(row=i,column=2,value=rw[1]).font=f(); ws.cell(row=i,column=3,value=rw[2]).font=f()
    c=ws.cell(row=i,column=4,value=rw[3]); c.number_format=EUR; c.fill=fill(BLUE_LT); c.font=f(color=BLUE_DK)
    ws.cell(row=i,column=5,value=rw[4]).font=f(); ws.cell(row=i,column=6,value=rw[5]).font=f()
GR_LAST=GR0+len(gastos)-1
zebra(ws,GR0,GR0+40-1,len(GH)); ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:F{GR0+40-1}"
cats='"Marketing,Material escritório,Equipamento,Deslocações,Taxas/bancos,Formação,Comunicações,Manutenção,Outros"'
dv=DataValidation(type="list",formula1=cats,allow_blank=True); ws.add_data_validation(dv); dv.add(f"B{GR0}:B1000")
dv=DataValidation(type="list",formula1='"Numerário,MB Way,Transferência,Cartão,Débito direto,Cheque,Outro"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"E{GR0}:E1000")
ws.cell(row=3,column=8,value="Por categoria").font=f(10,b=True,color=PINK)
catlist=["Marketing","Material escritório","Equipamento","Deslocações","Taxas/bancos","Formação","Comunicações","Manutenção","Outros"]
for k,cat in enumerate(catlist):
    ws.cell(row=4+k,column=8,value=cat).font=f(9)
    c=ws.cell(row=4+k,column=9,value=f'=SUMIF($B${GR0}:$B$1000,$H{4+k},$D${GR0}:$D$1000)'); c.number_format=EUR; c.font=f(9)
gtr=4+len(catlist)
ws.cell(row=gtr,column=8,value="TOTAL").font=f(10,b=True,color=PINK)
c=ws.cell(row=gtr,column=9,value=f'=SUM($D${GR0}:$D$1000)'); c.number_format=EUR; c.font=f(10,b=True,color=PINK)
ws.column_dimensions["G"].width=3; ws.column_dimensions["H"].width=18; ws.column_dimensions["I"].width=12
GA="'Gastos Adicionais'"

# ===== FLUXO DE CAIXA =====
ws=wb.create_sheet("Fluxo de Caixa"); nogrid(ws); tab(ws,BLUE_MD)
ws["A1"]="🔁 Fluxo de caixa"; ws["A1"].font=Font(name=FONT,bold=True,size=15,color=BLUE)
H=["Data","Descrição","Categoria","Entrada (€)","Saída (€)","Saldo acumulado (€)"]; W=[12,32,22,14,14,18]
paint_header(ws,2,H,W,fillhex=BLUE_MD)
mov=[["2026-02-01","Saldo inicial","Abertura",2000,0],["2026-02-12","Sinal R-2026-001 (João Silva)","Recebimento cliente",430,0],
 ["2026-02-20","Pagamento R-2026-002 (Maria Santos)","Recebimento cliente",8400,0],["2026-02-22","Voos+hotel Maldivas","Pagamento fornecedor",0,6900],
 ["2026-02-28","Custos fixos de fevereiro","Custo fixo",0,2592],["2026-03-05","Pagamento R-2026-003 (Pedro Costa)","Recebimento cliente",1250,0],
 ["2026-03-06","Voos+hotel Roma","Pagamento fornecedor",0,890],["2026-03-15","Pagamento R-2026-004 (Ana Ferreira)","Recebimento cliente",145,0],
 ["2026-03-22","Prestações Safari R-2026-005","Recebimento cliente",5200,0],["2026-03-28","Comissões de fevereiro","Comissão",0,102],
 ["2026-03-31","Custos fixos de março","Custo fixo",0,2592]]
r0=3
for i,rw in enumerate(mov,r0):
    c=ws.cell(row=i,column=1,value=D(rw[0])); c.number_format=DATE; c.font=f()
    ws.cell(row=i,column=2,value=rw[1]).font=f(); ws.cell(row=i,column=3,value=rw[2]).font=f()
    c=ws.cell(row=i,column=4,value=rw[3]); c.number_format=EUR; c.fill=fill(BLUE_LT); c.font=f(color=BLUE_DK)
    c=ws.cell(row=i,column=5,value=rw[4]); c.number_format=EUR; c.fill=fill(BLUE_LT); c.font=f(color=BLUE_DK)
    c=ws.cell(row=i,column=6,value=(f'=$D{i}-$E{i}' if i==r0 else f'=$F{i-1}+$D{i}-$E{i}')); c.number_format=EUR; c.font=f(b=True)
    for col in range(1,7): ws.cell(row=i,column=col).border=BOT
rN=r0+len(mov)-1; BALm=f"$F${rN}"; ws.freeze_panes="A3"
# --- auto: gastos adicionais (da aba «Gastos Adicionais») como saídas ---
gsr=rN+2
ws.cell(row=gsr,column=2,value="— Gastos adicionais (automático) —").font=f(9,it=True,color=PINK)
gsr+=1
FEED=30
for k in range(FEED):
    rr=gsr+k; gr=GR0+k; gg=f'{GA}!$A{gr}=""'
    c=ws.cell(row=rr,column=1,value=f'=IF({gg},"",{GA}!$A{gr})'); c.number_format=DATE; c.font=f()
    ws.cell(row=rr,column=2,value=f'=IF({gg},"","Gasto: "&{GA}!$C{gr})').font=f()
    ws.cell(row=rr,column=3,value=f'=IF({gg},"",{GA}!$B{gr})').font=f()
    c=ws.cell(row=rr,column=5,value=f'=IF({gg},"",{GA}!$D{gr})'); c.number_format=EUR
    c=ws.cell(row=rr,column=6,value=f'=IF({gg},"",{BALm}-SUM({GA}!$D${GR0}:$D{gr}))'); c.number_format=EUR; c.font=f(b=True)
    for col in range(1,7): ws.cell(row=rr,column=col).border=BOT
glast=gsr+FEED-1
ws.cell(row=glast+2,column=5,value="Saldo de caixa (final):").font=f(10,b=True); ws.cell(row=glast+2,column=5).alignment=Alignment(horizontal="right")
c=ws.cell(row=glast+2,column=6,value=f'={BALm}-SUM({GA}!$D${GR0}:$D$1000)'); c.number_format=EUR; c.font=f(b=True,color=BLUE)
SALDO=f"'Fluxo de Caixa'!$F${glast+2}"

# ===== CONTAS A RECEBER =====
ws=wb.create_sheet("Contas a Receber"); nogrid(ws); tab(ws,ORANGE)
ws["A1"]="⏳ Contas a receber"; ws["A1"].font=Font(name=FONT,bold=True,size=15,color=BLUE)
ws["A2"]="Quanto falta receber de cada cliente"; ws["A2"].font=f(9,it=True,color=MUTE)
H=["ID Reserva","Cliente","Viagem","Total (€)","Recebido (€)","Em falta (€)","Prestações","Faltam","Próx. prestação (€)","% recebido","Progresso"]
W=[13,17,20,13,14,14,11,8,17,11,16]; paint_header(ws,3,H,W,fillhex=ORANGE)
NR=40
for k in range(NR):
    rr=4+k; src=3+k; g=f'{IR}!$A{src}=""'
    ws.cell(row=rr,column=1,value=f'=IF({g},"",{IR}!$A{src})')
    ws.cell(row=rr,column=2,value=f'=IF({g},"",{IR}!$D{src})')
    ws.cell(row=rr,column=3,value=f'=IF({g},"",{IR}!$F{src})')
    c=ws.cell(row=rr,column=4,value=f'=IF({g},"",{IR}!$O{src})'); c.number_format=EUR
    c=ws.cell(row=rr,column=5,value=f'=IF({g},"",{IR}!$V{src})'); c.number_format=EUR
    c=ws.cell(row=rr,column=6,value=f'=IF({g},"",{IR}!$W{src})'); c.number_format=EUR
    ws.cell(row=rr,column=7,value=f'=IF({g},"",{IR}!$T{src}&" / "&{IR}!$S{src})')
    ws.cell(row=rr,column=8,value=f'=IF({g},"",{IR}!$X{src})')
    c=ws.cell(row=rr,column=9,value=f'=IF({g},"",IF({IR}!$X{src}>0,{IR}!$U{src},0))'); c.number_format=EUR
    c=ws.cell(row=rr,column=10,value=f'=IF({g},"",{IR}!$Y{src})'); c.number_format=PCT
    ws.cell(row=rr,column=11,value=(f'=IF({g},"",REPT("█",ROUND(IF({IR}!$Y{src}="",0,{IR}!$Y{src})*10,0))'
                                    f'&REPT("░",10-ROUND(IF({IR}!$Y{src}="",0,{IR}!$Y{src})*10,0)))'))
zebra(ws,4,4+NR-1,len(H))
for k in range(NR): ws.cell(row=4+k,column=11).font=Font(name=FONT,size=11,color=ORANGE)
ws.freeze_panes="A4"
ar=4+NR+2
ws.cell(row=ar,column=2,value="Já recebido").font=f(10,b=True)
c=ws.cell(row=ar,column=3,value=f"=SUM(E4:E{3+NR})"); c.number_format=EUR; c.font=f(11,b=True,color=BLUE)
ws.cell(row=ar+1,column=2,value="A receber").font=f(10,b=True)
c=ws.cell(row=ar+1,column=3,value=f"=SUM(F4:F{3+NR})"); c.number_format=EUR; c.font=f(11,b=True,color=ORANGE)
AR_TOT=f"'Contas a Receber'!$C${ar+1}"; REC_TOT=f"'Contas a Receber'!$C${ar}"
pie=PieChart(); pie.title="Recebido vs. a receber"
pie.add_data(Reference(ws,min_col=3,min_row=ar,max_row=ar+1)); pie.set_categories(Reference(ws,min_col=2,min_row=ar,max_row=ar+1))
pie.height=6.5; pie.width=9.5; ws.add_chart(pie,"E"+str(ar-1))

# ===== RENDIMENTOS E LUCROS =====
ws=wb.create_sheet("Rendimentos e Lucros"); nogrid(ws); tab(ws,BLUE)
ws["A1"]="📈 Rendimentos e lucros (P&L)"; ws["A1"].font=Font(name=FONT,bold=True,size=15,color=BLUE)
ws["A2"]="Por mês, com base na data de registo"; ws["A2"].font=f(9,it=True,color=MUTE)
months=[("fev/26",datetime(2026,2,1),datetime(2026,2,28)),("mar/26",datetime(2026,3,1),datetime(2026,3,31)),("abr/26",datetime(2026,4,1),datetime(2026,4,30))]
paint_header(ws,4,["Rubrica"]+[m[0] for m in months]+["Total"],[26,14,14,14,14])
def sm(col,mi): return f'SUMIFS({IR}!${col}:${col},{IR}!$B:$B,">="&DATE({months[mi][1].year},{months[mi][1].month},1),{IR}!$B:$B,"<="&DATE({months[mi][2].year},{months[mi][2].month},{months[mi][2].day}))'
def cm(mi): return f'SUMIFS(Comissões!$H$11:$H$42,Comissões!$B$11:$B$42,">="&DATE({months[mi][1].year},{months[mi][1].month},1),Comissões!$B$11:$B$42,"<="&DATE({months[mi][2].year},{months[mi][2].month},{months[mi][2].day}))'
rub=["Receita (venda)","Custo dos fornecedores","Margem bruta","Comissões","Custos fixos (mensais)","Gastos adicionais","Lucro líquido"]; base=5
for j,label in enumerate(rub):
    c=ws.cell(row=base+j,column=1,value=label); c.font=f(b=(label in("Margem bruta","Lucro líquido"))); c.border=BOT
def gm(mi): return f'SUMIFS({GA}!$D:$D,{GA}!$A:$A,">="&DATE({months[mi][1].year},{months[mi][1].month},1),{GA}!$A:$A,"<="&DATE({months[mi][2].year},{months[mi][2].month},{months[mi][2].day}))'
def scusto(mi): return f'SUMIFS({IR}!$N:$N,{IR}!$B:$B,">="&DATE({months[mi][1].year},{months[mi][1].month},1),{IR}!$B:$B,"<="&DATE({months[mi][2].year},{months[mi][2].month},{months[mi][2].day}),{IR}!$O:$O,">0")'
for mi in range(len(months)):
    col=2+mi; L=get_column_letter(col)
    ws.cell(row=base,column=col,value=f'={sm("O",mi)}')
    ws.cell(row=base+1,column=col,value=f'={scusto(mi)}')
    ws.cell(row=base+2,column=col,value=f'=${L}{base}-${L}{base+1}').font=f(b=True)
    ws.cell(row=base+3,column=col,value=f'={cm(mi)}')
    ws.cell(row=base+4,column=col,value=f'={CF}')
    ws.cell(row=base+5,column=col,value=f'={gm(mi)}')
    ws.cell(row=base+6,column=col,value=f'=${L}{base+2}-${L}{base+3}-${L}{base+4}-${L}{base+5}').font=f(b=True,color="375623")
    for rr in range(base,base+7): ws.cell(row=rr,column=col).number_format=EUR; ws.cell(row=rr,column=col).border=BOT
tcol=2+len(months); TL=get_column_letter(tcol); first=get_column_letter(2); last=get_column_letter(1+len(months))
for j in range(7):
    rr=base+j; c=ws.cell(row=rr,column=tcol,value=f'=SUM(${first}{rr}:${last}{rr})'); c.number_format=EUR; c.border=BOT
    c.font=f(b=True,color=("375623" if j==6 else BLUE if j==2 else TEXT))
RECEITA=f"'Rendimentos e Lucros'!${TL}${base}"; CUSTO=f"'Rendimentos e Lucros'!${TL}${base+1}"
MARGEM=f"'Rendimentos e Lucros'!${TL}${base+2}"; GASTOS=f"'Rendimentos e Lucros'!${TL}${base+5}"; LUCRO=f"'Rendimentos e Lucros'!${TL}${base+6}"

# ===== DASHBOARD =====
ws=wb.create_sheet("Dashboard"); nogrid(ws); tab(ws,BLUE_DK)
ws["B2"]="📊 Dashboard financeiro"; ws["B2"].font=Font(name=FONT,bold=True,size=18,color=BLUE)
ws["B3"]="Let's Go · confidencial"; ws["B3"].font=f(10,it=True,color=PINK)
# --- tabelas auxiliares (linhas 40+) para «mais vendido» ---
pac_names=list(dict.fromkeys(PACN.values())); destinos=list(dict.fromkeys(PACD.values()))
HP=40
for k,nm in enumerate(pac_names):
    ws.cell(row=HP+k,column=12,value=nm); ws.cell(row=HP+k,column=13,value=f'=COUNTIF({IR}!$F$3:$F$42,$L{HP+k})')
for k,nm in enumerate(destinos):
    ws.cell(row=HP+k,column=15,value=nm); ws.cell(row=HP+k,column=16,value=f'=COUNTIF({IR}!$H$3:$H$42,$O{HP+k})')
np=len(pac_names); nd=len(destinos)
Pmax=f'IFERROR(INDEX($L${HP}:$L${HP+np-1},MATCH(MAX($M${HP}:$M${HP+np-1}),$M${HP}:$M${HP+np-1},0)),"")'
Dmax=f'IFERROR(INDEX($O${HP}:$O${HP+nd-1},MATCH(MAX($P${HP}:$P${HP+nd-1}),$P${HP}:$P${HP+nd-1},0)),"")'
# --- KPIs numéricos (3 por linha) ---
RESC=f"COUNTA({IR}!$A$3:$A$42)"
NCLI=f'SUMPRODUCT(({IR}!$D$3:$D$42<>"")/COUNTIF({IR}!$D$3:$D$42,{IR}!$D$3:$D$42&""))'
MAXD=f"MAX({IR}!$B$3:$B$42)"
RECREF=f'SUMIFS({IR}!$O$3:$O$42,{IR}!$B$3:$B$42,">="&EOMONTH({MAXD},-1)+1,{IR}!$B$3:$B$42,"<="&EOMONTH({MAXD},0))'
RECANT=f'SUMIFS({IR}!$O$3:$O$42,{IR}!$B$3:$B$42,">="&EOMONTH({MAXD},-2)+1,{IR}!$B$3:$B$42,"<="&EOMONTH({MAXD},-1))'
kpis=[("Faturação total",RECEITA,EUR,BLUE),("Margem bruta",MARGEM,EUR,BLUE),("Lucro previsto (faturado)",LUCRO,EUR,ORANGE),
 ("Lucro realizado (caixa)",f"{LUCRO}-{AR_TOT}",EUR,ORANGE),("A receber",AR_TOT,EUR,ORANGE),("Já recebido",REC_TOT,EUR,BLUE),
 ("Comissões a pagar",COMM_TOTAL,EUR,BLUE),("Custos fixos (mês)",CF,EUR,BLUE),("Gastos adicionais",GASTOS,EUR,ORANGE),
 ("Saldo de caixa",SALDO,EUR,BLUE),("Nº de reservas",RESC,'0',BLUE_MD),("Nº de clientes",NCLI,'0',BLUE_MD),
 ("Ticket médio / cliente",f"IFERROR({RECEITA}/({NCLI}),0)",EUR,BLUE_MD),("Ticket médio / reserva",f"IFERROR({RECEITA}/{RESC},0)",EUR,BLUE_MD),
 ("Receita do último mês",RECREF,EUR,BLUE),("Variação vs mês anterior",f"IFERROR(({RECREF}-({RECANT}))/({RECANT}),0)",'0.0%',ORANGE)]
sr=5
for idx,(label,expr,fmt,col) in enumerate(kpis):
    cc=2+(idx%3)*3; rr=sr+(idx//3)*3; accent=fill(ORANGE_LT) if col==ORANGE else fill(BLUE_LT)
    b=ws.cell(row=rr,column=cc,value=label); b.font=f(9,b=True,color=MUTE); b.fill=accent; b.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells(start_row=rr,start_column=cc,end_row=rr,end_column=cc+1)
    v=ws.cell(row=rr+1,column=cc,value="="+expr); v.number_format=fmt; v.font=Font(name=FONT,bold=True,size=15,color=col); v.fill=accent
    v.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells(start_row=rr+1,start_column=cc,end_row=rr+1,end_column=cc+1)
    ws.row_dimensions[rr].height=15; ws.row_dimensions[rr+1].height=23
for cL,wd in [("A",2),("B",15),("C",15),("D",3),("E",15),("F",15),("G",3),("H",15),("I",15)]: ws.column_dimensions[cL].width=wd
# --- Destaques (texto) ---
dr0=sr+((len(kpis)+2)//3)*3+1
ws.cell(row=dr0,column=2,value="Destaques").font=f(12,b=True,color=BLUE_DK)
for k,(lab,ex) in enumerate([("Pacote mais vendido",Pmax),("Destino mais vendido",Dmax)]):
    rr=dr0+1+k
    l=ws.cell(row=rr,column=2,value=lab); l.font=f(10,b=True,color=BLUE_DK)
    v=ws.cell(row=rr,column=4,value="="+ex); v.font=f(10,color=PINK); v.fill=fill(BLUE_ZEBRA); v.border=BOT
    ws.merge_cells(start_row=rr,start_column=4,end_row=rr,end_column=6)
# --- gráficos ---
ch0=dr0+5
aux=46
for k,(lbl,frm) in enumerate([("Receita",RECEITA),("Custo",CUSTO),("Margem",MARGEM),("Lucro",LUCRO)]):
    ws.cell(row=aux+k,column=12,value=lbl); c=ws.cell(row=aux+k,column=13,value=f'={frm}'); c.number_format=EUR
bar=BarChart(); bar.title="Receita · Custo · Margem · Lucro"; bar.type="col"; bar.legend=None
bar.add_data(Reference(ws,min_col=13,min_row=aux,max_row=aux+3)); bar.set_categories(Reference(ws,min_col=12,min_row=aux,max_row=aux+3))
bar.height=7; bar.width=12; ws.add_chart(bar,f"B{ch0}")
catr=52
for k,cat in enumerate(["Instalações","Salários","Software","Serviços"]):
    ws.cell(row=catr+k,column=12,value=cat); c=ws.cell(row=catr+k,column=13,value=f"=SUMIF('Custos Fixos'!$B:$B,$L{catr+k},'Custos Fixos'!$G:$G)"); c.number_format=EUR
pie=PieChart(); pie.title="Custos fixos por categoria"
pie.add_data(Reference(ws,min_col=13,min_row=catr,max_row=catr+3)); pie.set_categories(Reference(ws,min_col=12,min_row=catr,max_row=catr+3))
pie.height=7; pie.width=10; ws.add_chart(pie,f"F{ch0}")

wb.save("/home/claude/Financeiro_Agencia.xlsx"); print("fin ok")
